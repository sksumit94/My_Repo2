from statistics import mode
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from robot.api.deco import keyword, library
from robot.api import logger
import  pandas as pd
import openpyxl
import string
#import xlsxwriter


@library(scope='TEST CASE', version='1.0.0')

 
class UpdateExcelCell:
    @keyword
    def color_cell_of_excel_with_Tcode_Status(self,filepath,row,col,status,flag):
        r'''
        Color the cell of the excel as per the matrix.
        '''
        sheetno = 2
        workbook = load_workbook(filepath)
        sheet = workbook.sheetnames
        worksheet = workbook[sheet[sheetno]]
    
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
            
        if status == True and flag == 'Y':
            fill_cell = PatternFill(patternType='solid', 
                            fgColor='40ff00')
            worksheet.cell(row=row,column=col).fill = fill_cell 
            worksheet.cell(row=row,column=col).border = border
        elif status == False and flag == 'Y':
            fill_cell = PatternFill(patternType='solid', 
                            fgColor='ff0000')
            worksheet.cell(row=row,column=col).fill = fill_cell 
            worksheet.cell(row=row,column=col).border = border
        elif status == True and flag == 'N':
            fill_cell = PatternFill(patternType='solid', 
                            fgColor='40ff00')
            worksheet.cell(row=row,column=col).fill = fill_cell 
            worksheet.cell(row=row,column=col).border = border
        elif status == False and flag == 'N':
            fill_cell = PatternFill(patternType='solid', 
                            fgColor='ff0000')
            worksheet.cell(row=row,column=col).fill = fill_cell
            worksheet.cell(row=row,column=col).border = border 
        else:
            logger.console('Verification for Tcode Security Matrix is not Present')
              
        workbook.save(filepath)
        workbook.close()

    @keyword
    def color_cell_for_tiles_FIORI_matrix(self,filepath,role,screenName,tileName,tile_access,issue_desc):
        r'''
        Color the cell of the excel as per the matrix.
        '''
        workbook = load_workbook(filepath)
        sheet = workbook.sheetnames
        sheetno = sheet.index(role)
        worksheet = workbook[sheet[sheetno]]
        for col in worksheet.iter_cols():
            for cell in col:
                if cell.value == 'ScreenName':
                    colIndex_ScreenName = cell.column
                if cell.value == 'TileName':
                    colIndex_TileName = cell.column
        rowCount = worksheet.max_row
        print(rowCount)
        print(screenName)
        print(tileName)
        for i in range(1,rowCount+1):
            cellVal_tileName = worksheet.cell(i,colIndex_TileName).value    
            cellVal_ScreenName = worksheet.cell(i,colIndex_ScreenName).value   
            print(cellVal_tileName) 
            print(cellVal_ScreenName) 
            if cellVal_tileName == tileName and cellVal_ScreenName == screenName:
                    rowIndex = i
                    col = colIndex_TileName
        print(rowIndex)
        print(col)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        if tile_access == True:
            fill_cell = PatternFill(patternType='solid', 
                            fgColor='40ff00')
            worksheet.cell(row=rowIndex,column=col).fill = fill_cell 
            worksheet.cell(row=rowIndex,column=col).border = border
        elif tile_access == False:
            fill_cell = PatternFill(patternType='solid', 
                            fgColor='ff0000')
            worksheet.cell(row=rowIndex,column=col).fill = fill_cell 
            worksheet.cell(row=rowIndex,column=col).border = border 
            col = col + 1
            worksheet.cell(row=rowIndex,column=col).value = issue_desc       
        else:
            logger.console('Verification for FIORI Security Matrix is not Present')
              
        workbook.save(filepath)
        workbook.close()

    @keyword
    def log_screen_level_error_msg(self,filepath,role,screen_name,tile_list_status,issue_desc):
        r'''
        Color the cell of the excel as per the matrix.
        '''
        workbook = load_workbook(filepath)
        sheet = workbook.sheetnames
        sheetno = sheet.index(role)
        worksheet = workbook[sheet[sheetno]]
        for col in worksheet.iter_cols():
            for cell in col:
                if cell.value == 'ScreenLevelErrorDesc':
                    colIndex = cell.column
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == screen_name:
                    rowIndex = cell.row

        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        if tile_list_status == True:
            fill_cell = PatternFill(patternType='solid', 
                            fgColor='40ff00')
            worksheet.cell(row=rowIndex,column=colIndex).fill = fill_cell 
            worksheet.cell(row=rowIndex,column=colIndex).border = border
            worksheet.cell(row=rowIndex,column=colIndex).value = issue_desc       
        elif tile_list_status == False:
            fill_cell = PatternFill(patternType='solid', 
                            fgColor='ff0000')
            worksheet.cell(row=rowIndex,column=colIndex).fill = fill_cell 
            worksheet.cell(row=rowIndex,column=colIndex).border = border 
            worksheet.cell(row=rowIndex,column=colIndex).value = issue_desc       
        else:
            logger.console('Verification for FIORI Security Matrix is not Present')
              
        workbook.save(filepath)
        workbook.close()

    @keyword
    def log_additional_error_msg(self,filepath,role,issue_desc):
        r'''
        Color the cell of the excel as per the matrix.
        '''
        workbook = load_workbook(filepath)
        sheet = workbook.sheetnames
        sheetno = sheet.index(role)
        worksheet = workbook[sheet[sheetno]]
        for col in worksheet.iter_cols():
            for cell in col:
                if cell.value == 'Additional Error Msg':
                    colIndex = cell.column
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == 'Additional Error Msg':
                    rowIndex = cell.row

        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        
        fill_cell = PatternFill(patternType='solid', 
                            fgColor='ff0000')
        worksheet.cell(row=rowIndex,column=colIndex + 1).fill = fill_cell 
        worksheet.cell(row=rowIndex,column=colIndex + 1).border = border 
        worksheet.cell(row=rowIndex,column=colIndex + 1).value = issue_desc                     
        workbook.save(filepath)
        workbook.close()
    
    @keyword
    def update_value_for_variable_excel(self,filepath,sheet_name,variable_name,variable_value):
        r'''
        Color the cell of the excel as per the matrix.
        '''
        workbook = load_workbook(filepath,data_only=True)
        sheet = workbook.sheetnames
        sheetno = sheet.index(sheet_name)
        worksheet = workbook[sheet[sheetno]]
        print ('start')
        for col in worksheet.iter_cols():
            for cell in col:
                if cell.value == variable_name:
                    print (variable_name)
                    colIndex = cell.column
        worksheet.cell(row=2,column=colIndex).value = variable_value                     
        workbook.save(filepath)
        workbook.close()
    
    @keyword
    def Create_output_excel(self,filepath,colheader):
        r'''
        Color the cell of the excel as per the matrix.
        '''
        workbook = openpyxl.Workbook()
        workbook.save(filepath)
        worksheet = workbook.active
        worksheet.title = "Output"
        i = 1
        colheader = list(colheader.split(","))
        for item in colheader:
            workbookcell = worksheet.cell(row=1,column=i)
            workbookcell.value = item
            font = Font(bold=True)
            workbookcell.font = font
            i = i+1
        workbook.save(filepath)
        print ('saved')
        workbook.close()

    @keyword
    def Update_Status_With_Testcase_Name(self,filepath,role,testcase_result):
        result_df = pd.DataFrame(list(testcase_result.items()),columns=['Testcase Name','Testcase Status'])
        # workbook = load_workbook(filepath)
        # with pd.ExcelWriter(filepath,mode='a') as writer:  
        #     result_df.to_excel(writer)
        writer = pd.ExcelWriter(filepath, engine='openpyxl',mode='a')
        # writer.book = workbook
        result_df.to_excel(writer,sheet_name=role)
        # workbook.save(filepath)
        # writer.save()
        writer.close()
        print("Data is saved succesfully in excel file")

    @keyword
    def Update_Status_In_Excel(self,filepath,sheetName,col_header,testcase_result):
        workbook = load_workbook(filepath) 
        sheet = workbook.sheetnames
        sheetno = sheet.index(sheetName)
        worksheet = workbook[sheet[sheetno]]
        colIndex = worksheet.max_column + 1
        result_df = pd.DataFrame(testcase_result,columns=[col_header])
        writer = pd.ExcelWriter(filepath, engine='openpyxl',mode='a',if_sheet_exists='overlay')
        result_df.to_excel(writer,sheet_name=sheetName,index=False,startcol=worksheet.max_column)
        # writer.save()
        writer.close()
        print("Data is saved succesfully in excel file")    
    
    @keyword
    def get_max_row_and_max_column_values(self,filepath,sheetName,valueType):
        '''if you pass valueType is row, column and rowandColumn then it returns maxrow,maxColumn and maxRowandColumn values repectively'''
        wb = load_workbook(filepath) 
        sheetName= str(sheetName) 
        sheet= wb[sheetName]
        print("{}rows and {}columns present in {}-{}".format(sheet.max_row,sheet.max_column,filepath,sheetName))
        if valueType=='row':
            return sheet.max_row
        elif valueType=='column':
            return sheet.max_column
        else:
            return list((sheet.max_row,sheet.max_column))
    
    # @keyword
    # def write_data_into_excel_by_column_name(self,filepath,sheetName,columnName,RowNum,data):
    #     '''args were filepath,columnName,RowNumber,value
    #     columnindex return -1 of value so while writing incrementing to +1'''
    #     wrkBk = pd.read_excel(filepath)
    #     columns = wrkBk.columns
    #     column_index = columns.get_loc(columnName)
    #     print(column_index)
    #     wb = load_workbook(filepath) 
    #     sheetName= str(sheetName) 
    #     sheet= wb[sheetName]
    #     print(sheet)
    #     if type(data)==int:
    #         sheet.cell(row=RowNum, column=int(column_index)+1).value = int(data)
    #     else:
    #         sheet.cell(row=RowNum, column=int(column_index)+1).value = str(data)
    #     wb.save(filepath)  
    @keyword
    def remove_empty_lines_in_the_text_file(self,filepath):
        '''removes the empty linse and strip the file data in each line'''
        with open(filepath,'r') as text:
            data = text.readlines()
        data = [line.replace(" ",'') for line in data if line.strip() != '']
        with open(filepath,'w') as text:
            text.writelines(data)
    